"""Exporta os pedidos selecionados em .xlsx formatado, uma aba por fornecedor.

Estrutura de cada aba:
  - título: "Fornecedor: X — De DD/MM/AAAA a DD/MM/AAAA — pedidos selecionados"
  - cabeçalho: Pedido | Quantidade | Item | Número | Cor | Obs
  - uma linha por item, com o número do pedido na mesma linha (repete a cada
    item quando o pedido tem mais de um) — nada de pedido numa linha e item
    na linha de baixo
  - Obs = "Pedido com mais de um item" quando o pedido tem +1 linha de item
"""
from __future__ import annotations

import io
import re
from datetime import date, datetime
from itertools import groupby

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADERS = ("Pedido", "Quantidade", "Item", "Número", "Cor", "Obs")
_WIDTHS = (16, 11, 42, 12, 24, 28)
SEM_FORN = "Sem fornecedor"


def _fmt_d(v: str | date | datetime | None) -> str | None:
    if not v:
        return None
    if isinstance(v, str):
        try:
            v = datetime.fromisoformat(v.replace("Z", "+00:00"))
        except ValueError:
            return v[:10]
    return v.strftime("%d/%m/%Y")


def _periodo(date_from: str | None, date_to: str | None, orders: list[dict]) -> str:
    a, b = _fmt_d(date_from), _fmt_d(date_to)
    if not a or not b:
        stamps = [o.get("placed_at") for o in orders if o.get("placed_at")]
        ds = sorted(_fmt_d(s) for s in stamps if _fmt_d(s))
        if ds:
            a, b = a or ds[0], b or ds[-1]
    if a and b:
        return f"De {a} a {b}" if a != b else f"Em {a}"
    return "Todos os períodos"


def _safe_sheet(name: str, used: set[str]) -> str:
    clean = re.sub(r"[\[\]:*?/\\]", "-", name).strip()[:31] or "Fornecedor"
    base, i = clean, 2
    while clean.lower() in used:
        suf = f" ({i})"
        clean = base[: 31 - len(suf)] + suf
        i += 1
    used.add(clean.lower())
    return clean


def _cor_from_name(name: str) -> str:
    # "BOTA COTURNO 2187 CAFE" -> "CAFE" ; "TENIS 2085 NUDE ROSA" -> "NUDE ROSA"
    m = re.match(
        r"(?:BOTA\s+COTURNO|COTURNO|T[ÊE]NIS|TENIS|BOTA|SAPAT[ÊE]NIS|SAND[ÁA]LIA)\s+\S+\s+(.+)",
        (name or "").strip(),
        re.I,
    )
    return m.group(1).strip() if m else ""


def _cor_numero(item: dict) -> tuple[str, str]:
    attrs = item.get("variant_attrs") or {}
    cor = str(attrs.get("cor") or "").strip()
    numero = str(attrs.get("numero") or "").strip()
    label = str(item.get("variant_label") or "").strip()
    if not numero:
        m = re.search(r"\d{2,3}", label)
        numero = m.group(0) if m else label
    if not cor:
        cor = str(item.get("product_color") or "").strip() or _cor_from_name(item.get("name", ""))
        # se a numeração veio de "42 / Preto", a parte não-numérica pode ser a cor
        if not cor and "/" in label:
            cor = label.split("/", 1)[1].strip()
    return cor, numero


def build_supplier_xlsx(
    orders: list[dict], *, date_from: str | None = None, date_to: str | None = None
) -> bytes:
    by_supplier: dict[str, list[tuple[dict, dict]]] = {}
    multi_item: dict[str, bool] = {}
    for o in orders:
        multi_item[o["number"]] = len(o.get("items") or []) > 1
        for it in o.get("items") or []:
            forn = str(it.get("supplier") or "").strip() or SEM_FORN
            by_supplier.setdefault(forn, []).append((o, it))

    wb = Workbook()
    wb.remove(wb.active)  # remove a aba padrão

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor="111111")
    head_font = Font(bold=True, color="FFFFFF")
    title_font = Font(bold=True, size=12)
    order_font = Font(bold=True, color="444444")
    order_fill = PatternFill("solid", fgColor="EFEFEF")
    wrap = Alignment(wrap_text=True, vertical="top")
    center = Alignment(horizontal="center", vertical="top")

    periodo = _periodo(date_from, date_to, orders)
    used_names: set[str] = set()

    order_forn = [
        forn
        for forn in sorted(by_supplier, key=lambda s: (s == SEM_FORN, s.lower()))
    ]
    if not order_forn:
        ws = wb.create_sheet("Sem itens")
        ws.append(["Nenhum item nos pedidos selecionados."])

    for forn in order_forn:
        entries = sorted(by_supplier[forn], key=lambda e: e[0]["number"])
        pedidos_distintos = {o["number"] for o, _ in entries}
        ws = wb.create_sheet(_safe_sheet(forn, used_names))

        ws.append([f"Fornecedor: {forn} — {periodo} — {len(pedidos_distintos)} pedido(s) selecionado(s)"])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
        ws["A1"].font = title_font
        ws.append([])

        hrow = ws.max_row + 1
        ws.append(list(HEADERS))
        for c in ws[hrow]:
            c.font = head_font
            c.fill = head_fill
            c.border = border
            c.alignment = center

        shade = False
        for num, grp in groupby(entries, key=lambda e: e[0]["number"]):
            # sombreado alternado por pedido — só pra continuar dando pra
            # distinguir onde um pedido termina e o outro começa, já que o
            # número agora repete linha a linha (sem separador).
            shade = not shade
            for _o, it in grp:
                cor, numero = _cor_numero(it)
                obs = "Pedido com mais de um item" if multi_item[num] else ""
                # número do pedido na MESMA linha do item (repete a cada item
                # do mesmo pedido) — nada de linha separada só com o número.
                ws.append([num, int(it.get("quantity") or 0), it.get("name", ""), numero, cor, obs])
                row = ws.max_row
                for c in ws[row]:
                    c.border = border
                    c.alignment = wrap
                    if shade:
                        c.fill = order_fill
                ws[f"A{row}"].font = order_font
                ws[f"B{row}"].alignment = center
                ws[f"D{row}"].alignment = center

        for i, w in enumerate(_WIDTHS, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = f"A{hrow + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
