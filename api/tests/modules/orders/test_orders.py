"""Testes do módulo `orders` — numeração, snapshot, transições, lookup convidado."""
from __future__ import annotations

import pytest

ADDRESS = {
    "recipient_name": "Cliente Teste",
    "zip": "20040002",
    "street": "Av. Rio Branco",
    "number": "1",
    "district": "Centro",
    "city": "Rio de Janeiro",
    "state": "RJ",
}


@pytest.fixture
async def variant(client, admin_token, auth_headers):
    h = auth_headers(admin_token)
    cat = (await client.post("/api/admin/categories", json={"name": "C"}, headers=h)).json()
    p = (
        await client.post(
            "/api/admin/products",
            json={"name": "Item", "category_id": cat["id"], "price_cents": 7000, "status": "active"},
            headers=h,
        )
    ).json()
    await client.put(
        f"/api/admin/products/{p['id']}/option-types",
        json=[{"name": "T", "values": [{"value": "U"}]}],
        headers=h,
    )
    vid = (await client.get(f"/api/products/{p['slug']}")).json()["option_types"][0]["values"][0]["id"]
    v = (
        await client.post(
            f"/api/admin/products/{p['id']}/variants",
            json={"sku": "IT-U", "option_value_ids": [vid], "stock_qty": 5},
            headers=h,
        )
    ).json()
    return v["id"]


async def _order(client, variant, email="c@test.example", qty=1):
    await client.post("/api/cart/items", json={"variant_id": variant, "quantity": qty})
    r = await client.post(
        "/api/orders/checkout", json={"email": email, "shipping_address": ADDRESS}
    )
    assert r.status_code == 201, r.text
    return r.json()


@pytest.mark.asyncio
async def test_number_is_sequential_per_year(client, variant):
    o1 = await _order(client, variant)
    o2 = await _order(client, variant)
    y1, n1 = o1["number"].split("-")
    y2, n2 = o2["number"].split("-")
    assert y1 == y2
    assert int(n2) == int(n1) + 1


@pytest.mark.asyncio
async def test_snapshot_and_stock_decrement(client, variant):
    order = await _order(client, variant, qty=2)
    assert order["items"][0]["unit_price_cents"] == 7000
    assert order["items"][0]["total_cents"] == 14000
    assert order["grand_total_cents"] == 14000
    # 5 - 2 = 3 em estoque
    plist = (await client.get(f"/api/products/{ (await client.get('/api/products?category=c')).json()['items'][0]['slug'] }")).json()
    assert plist["variants"][0]["stock_qty"] == 3


@pytest.mark.asyncio
async def test_guest_lookup_requires_email(client, variant):
    order = await _order(client, variant, email="guest@test.example")
    ok = await client.get(f"/api/orders/{order['number']}", params={"email": "guest@test.example"})
    assert ok.status_code == 200
    bad = await client.get(f"/api/orders/{order['number']}", params={"email": "outro@test.example"})
    assert bad.status_code == 404
    # IDOR: sem e-mail nenhum não pode devolver o pedido (evita enumeração por número)
    no_email = await client.get(f"/api/orders/{order['number']}")
    assert no_email.status_code == 404
    pulse = await client.get(f"/api/orders/{order['number']}/pulse")
    assert pulse.status_code == 404


@pytest.mark.asyncio
async def test_admin_status_transitions(client, variant, admin_token, auth_headers):
    h = auth_headers(admin_token)
    order = await _order(client, variant)
    num = order["number"]

    # o admin pode pular para qualquer status (fluxo livre); status inexistente é 422
    invalid = await client.post(f"/api/admin/orders/{num}/status", json={"status": "voando"}, headers=h)
    assert invalid.status_code == 422
    jump = await client.post(f"/api/admin/orders/{num}/status", json={"status": "shipped"}, headers=h)
    assert jump.status_code == 200

    ok = await client.post(f"/api/admin/orders/{num}/status", json={"status": "paid"}, headers=h)
    assert ok.status_code == 200
    ok2 = await client.post(f"/api/admin/orders/{num}/status", json={"status": "processing"}, headers=h)
    assert ok2.json()["status"] == "processing"

    # cancelamento restaura estoque
    await client.post(f"/api/admin/orders/{num}/status", json={"status": "canceled"}, headers=h)
    detail = (await client.get(f"/api/admin/orders/{num}", headers=h)).json()
    assert detail["status"] == "canceled"


@pytest.mark.asyncio
async def test_cancel_restores_stock(client, variant, admin_token, auth_headers):
    h = auth_headers(admin_token)
    order = await _order(client, variant, qty=3)  # estoque 5 -> 2
    await client.post(f"/api/admin/orders/{order['number']}/status", json={"status": "canceled"}, headers=h)
    slug = (await client.get("/api/products?category=c")).json()["items"][0]["slug"]
    assert (await client.get(f"/api/products/{slug}")).json()["variants"][0]["stock_qty"] == 5


@pytest.mark.asyncio
async def test_checkout_records_processing_error_on_failed_step(client, variant, monkeypatch):
    """Se um passo pós-pedido falha, o pedido é criado igual e o motivo fica em
    `processing_error` (visível na conta / pulse)."""
    from app.modules.orders import events

    async def _boom(db, order):  # noqa: ANN001
        raise RuntimeError("smtp caiu")

    monkeypatch.setattr(events, "_send_account_access", _boom)

    order = await _order(client, variant, email="err@test.example")
    # a finalização roda em BackgroundTask; no TestClient ela já concluiu aqui
    pulse = await client.get(f"/api/orders/{order['number']}/pulse", params={"email": "err@test.example"})
    assert pulse.status_code == 200
    assert "dados de acesso" in (pulse.json().get("processing_error") or "")
    assert "smtp caiu" in pulse.json()["processing_error"]


@pytest.mark.asyncio
async def test_checkout_no_processing_error_on_happy_path(client, variant):
    order = await _order(client, variant, email="ok@test.example")
    pulse = await client.get(f"/api/orders/{order['number']}/pulse", params={"email": "ok@test.example"})
    assert pulse.json().get("processing_error") is None


@pytest.mark.asyncio
async def test_supplier_xlsx_export(client, variant, admin_token, auth_headers):
    h = auth_headers(admin_token)
    o1 = await _order(client, variant)
    o2 = await _order(client, variant)
    r = await client.get(
        "/api/admin/orders/export/suppliers.xlsx",
        params={"numbers": f"{o1['number']},{o2['number']}", "from": "2026-09-01", "to": "2026-09-30"},
        headers=h,
    )
    assert r.status_code == 200, r.text
    assert r.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert r.content[:2] == b"PK"  # zip container do .xlsx

    import io

    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(r.content))
    assert wb.sheetnames  # ao menos uma aba (por fornecedor)
    ws = wb[wb.sheetnames[0]]
    rows = [tuple(c for c in row) for row in ws.iter_rows(values_only=True)]
    assert any(row and str(row[0]).startswith("Fornecedor:") for row in rows)
    assert ("Pedido", "Quantidade", "Item", "Número", "Cor", "Obs") in rows
